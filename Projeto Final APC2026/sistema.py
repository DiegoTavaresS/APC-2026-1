# -*- coding: utf-8 -*-
"""
Sistema PDAD 2024 — Recorte E: Trabalho e Ocupação
Execute com: python sistema.py
Dependências: pandas, matplotlib, openpyxl
"""

import glob
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
import matplotlib
try:
    matplotlib.use("TkAgg")
except Exception:
    pass
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
SENTINELAS = {99999, 88888, 77777}

COLUNAS = [
    "A01nficha", "localidade", "idade_calculada", "id_genero", "escolaridade",
    "I01", "I04", "I13", "I20", "I20_r"
]

MAPA_OCUPACAO = {
    1: "Empregado no setor público", 2: "Militar", 3: "Empregado no setor privado",
    4: "Empregado doméstico", 5: "Estágio remunerado", 6: "Aprendiz",
    7: "Conta própria/autônomo", 8: "Empregador", 9: "Serviço militar obrigatório",
    10: "Trabalhador não remunerado"
}
MAPA_GENERO = {1: "Cisgênero", 2: "Transgênero", 3: "Outro"}
MAPA_ESCOLARIDADE = {
    1: "Sem instrução", 2: "Fundamental incompleto", 3: "Fundamental completo",
    4: "Médio incompleto", 5: "Médio completo", 6: "Superior incompleto",
    7: "Superior completo", 8: "Sem classificação"
}


def localizar_arquivo(padroes):
    """Procura um arquivo na pasta do sistema ou dentro da pasta dados."""
    base = os.path.dirname(os.path.abspath(__file__))
    for pasta in [base, os.path.join(base, "dados")]:
        for padrao in padroes:
            achados = glob.glob(os.path.join(pasta, padrao))
            if achados:
                return achados[0]
    return None


def numero_limpo(serie):
    """Converte coluna em número e substitui sentinelas por NA."""
    s = pd.to_numeric(serie.astype(str).str.replace(",", ".", regex=False), errors="coerce")
    return s.mask(s.isin(SENTINELAS))


def moeda(valor):
    """Formata número como dinheiro em reais."""
    if pd.isna(valor):
        return "R$ 0,00"
    texto = f"R$ {valor:,.2f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def rotulo_codigo(valor, mapa):
    """Mostra código e descrição lado a lado."""
    if pd.isna(valor):
        return "Sem informação"
    codigo = int(valor)
    return f"{codigo} - {mapa.get(codigo, 'Sem descrição')}"


def carregar_mapa_localidade(arquivo_dic):
    """Lê no dicionário o nome das localidades/Regiões Administrativas."""
    if not arquivo_dic:
        return {}
    try:
        anexo = pd.read_excel(arquivo_dic, sheet_name="anexo_1")
        anexo["Coluna"] = anexo["Coluna"].ffill().astype(str).str.strip()
        anexo = anexo[anexo["Coluna"].eq("localidade")].copy()
        anexo["Valor"] = pd.to_numeric(anexo["Valor"], errors="coerce")
        anexo = anexo.dropna(subset=["Valor", "Descrição do valor"])
        return {int(r["Valor"]): str(r["Descrição do valor"]).strip() for _, r in anexo.iterrows()}
    except Exception:
        return {}


def ordenar_manual(lista, indice=1, decrescente=True):
    """Ordena uma lista manualmente usando Selection Sort."""
    itens = list(lista)
    for i in range(len(itens)):
        melhor = i
        for j in range(i + 1, len(itens)):
            troca = itens[j][indice] > itens[melhor][indice] if decrescente else itens[j][indice] < itens[melhor][indice]
            if troca:
                melhor = j
        itens[i], itens[melhor] = itens[melhor], itens[i]
    return itens




def contar_linhas_excel(caminho):
    """Conta linhas de uma planilha sem carregar todos os dados no pandas."""
    if not caminho:
        return 0
    try:
        from openpyxl import load_workbook
        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.active
        total = max(ws.max_row - 1, 0)
        wb.close()
        return total
    except Exception:
        return len(pd.read_excel(caminho, usecols=[0]))


def avisar(progresso, texto, valor):
    """Atualiza a tela de carregamento quando ela existir."""
    if progresso:
        progresso(texto, valor)


def resumo_situacao(moradores):
    """Calcula ocupados, desempregados e fora da ocupação nos últimos 30 dias."""
    i04 = moradores["I04"]
    i01 = moradores["I01"]
    validos = moradores[i04.isin([1, 2])]
    total = len(validos)
    ocupados = len(validos[validos["I04"] == 1])
    desempregados = len(validos[(validos["I04"] == 2) & (validos["I01"] == 1)])
    fora = len(validos[(validos["I04"] == 2) & (validos["I01"] != 1)])
    pct = lambda n: (n / total * 100) if total else 0
    return ocupados, desempregados, fora, pct(ocupados), pct(desempregados), pct(fora)


def carregar_dados(progresso=None):
    """Carrega, limpa e prepara os dados usados pelo sistema."""
    avisar(progresso, "Procurando arquivos da PDAD...", 5)
    arq_mor = localizar_arquivo(["PDAD_2024-Moradores*.csv", "*Moradores*.csv"])
    arq_dom = localizar_arquivo(["PDAD_2024-Domicilios*.xlsx", "*Domicilios*.xlsx", "*Domicílios*.xlsx"])
    arq_dic = localizar_arquivo(["*dicionario*.xlsx", "*Dicionario*.xlsx", "*Dicionário*.xlsx"])

    if not arq_mor:
        raise FileNotFoundError("Coloque o arquivo PDAD_2024-Moradores.csv na pasta dados/ ou junto do sistema.py")

    avisar(progresso, "Lendo cabeçalho do arquivo de moradores...", 12)
    cabecalho = pd.read_csv(arq_mor, sep=";", nrows=0, encoding="utf-8-sig")
    cols = [c for c in COLUNAS if c in cabecalho.columns]

    avisar(progresso, "Carregando moradores. Isso pode demorar alguns segundos...", 25)
    moradores = pd.read_csv(arq_mor, sep=";", encoding="utf-8-sig", usecols=cols, low_memory=False)

    colunas_numericas = ["localidade", "idade_calculada", "id_genero", "escolaridade", "I01", "I04", "I13", "I20", "I20_r"]
    for i, col in enumerate(colunas_numericas):
        if col in moradores:
            avisar(progresso, f"Limpando coluna {col} e removendo sentinelas...", 35 + i * 3)
            moradores[col] = numero_limpo(moradores[col])

    avisar(progresso, "Lendo nomes das localidades no dicionário...", 65)
    localidades = carregar_mapa_localidade(arq_dic)

    avisar(progresso, "Criando rótulos de localidade, gênero, escolaridade e ocupação...", 75)
    moradores["ra"] = moradores["localidade"].apply(lambda x: rotulo_codigo(x, localidades))
    moradores["genero"] = moradores["id_genero"].apply(lambda x: rotulo_codigo(x, MAPA_GENERO))
    moradores["ensino"] = moradores["escolaridade"].apply(lambda x: rotulo_codigo(x, MAPA_ESCOLARIDADE))
    moradores["ocupacao"] = moradores["I13"].apply(lambda x: rotulo_codigo(x, MAPA_OCUPACAO))
    moradores["renda"] = moradores["I20_r"].fillna(moradores["I20"])

    avisar(progresso, "Separando pessoas ocupadas nos últimos 30 dias...", 85)
    ocupados = moradores[moradores["I04"] == 1].copy()

    avisar(progresso, "Contando domicílios...", 92)
    total_dom = contar_linhas_excel(arq_dom) if arq_dom else 0

    avisar(progresso, "Finalizando interface...", 100)
    return moradores, ocupados, total_dom


def estatisticas(df):
    """Calcula estatísticas principais do filtro atual."""
    rendas = df["renda"].dropna()
    return {
        "ocupados": len(df),
        "com_renda": len(rendas),
        "media": rendas.mean() if len(rendas) else 0,
        "mediana": rendas.median() if len(rendas) else 0,
    }


class SistemaPDAD:
    """Interface gráfica do sistema PDAD."""

    def __init__(self, root, moradores, ocupados, total_dom):
        self.root = root
        self.moradores = moradores
        self.ocupados = ocupados
        self.filtrado = ocupados.copy()
        self.total_dom = total_dom
        self.root.title("PDAD 2024 — Trabalho e Ocupação")
        self.root.geometry("1180x760")
        self.montar_interface()
        self.atualizar()

    def montar_interface(self):
        """Monta todos os widgets principais da janela."""
        topo = ttk.Frame(self.root, padding=12)
        topo.pack(fill="x")
        ttk.Label(topo, text="PDAD 2024 — Trabalho e Ocupação", font=("Segoe UI", 18, "bold")).pack(anchor="w")
        ttk.Label(topo, text="Análise do trabalho nos últimos 30 dias, com filtros por localidade, gênero e escolaridade.").pack(anchor="w")
        ttk.Label(topo, text="Sentinelas tratadas: 99999 = não se aplica / 88888 = nao declarado.").pack(anchor="w")
        ttk.Label(topo, text="Aluno: Diego Tavares Silva").pack(anchor="w")
        
        o, d, f, po, pd, pf = resumo_situacao(self.moradores)
        texto = (
            f"Moradores: {len(self.moradores):,} · Domicílios: {self.total_dom:,} · "
            f"Ocupados: {o:,} ({po:.1f}%) · Desempregados/procurando: {d:,} ({pd:.1f}%) · "
            f"Fora da ocupação: {f:,} ({pf:.1f}%)"
        ).replace(",", ".")
        ttk.Label(topo, text=texto, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(6, 0))

        filtros = ttk.LabelFrame(self.root, text="Filtros", padding=10)
        filtros.pack(fill="x", padx=12)
        self.var_ra = tk.StringVar(value="Todas")
        self.var_genero = tk.StringVar(value="Todos")
        self.var_ensino = tk.StringVar(value="Todos")
        self.var_renda = tk.StringVar(value="0")

        self.combo_ra = self.combo(filtros, "Localidade:", self.var_ra, self.opcoes("ra", "Todas"), 0)
        self.combo_genero = self.combo(filtros, "Gênero:", self.var_genero, self.opcoes("genero", "Todos"), 1)
        self.combo_ensino = self.combo(filtros, "Escolaridade:", self.var_ensino, self.opcoes("ensino", "Todos"), 2)
        self.combo_renda = self.combo(filtros, "Renda mínima:", self.var_renda, ["0", "1000", "2000", "5000", "10000"], 3)
        ttk.Button(filtros, text="Aplicar filtros", command=self.atualizar).grid(row=0, column=8, padx=8)
        ttk.Button(filtros, text="Gerar relatório TXT", command=self.gerar_relatorio).grid(row=0, column=9, padx=8)

        # Cards antigos de resumo: aparecem sempre no topo e mudam com os filtros.
        self.cards = ttk.Frame(self.root, padding=(12, 10, 12, 0))
        self.cards.pack(fill="x")
        self.card_vars = {}
        for titulo in ["Ocupados filtrados", "Com renda válida", "Média da renda", "Mediana da renda"]:
            card = ttk.LabelFrame(self.cards, text=titulo, padding=10)
            card.pack(side="left", fill="x", expand=True, padx=4)
            var = tk.StringVar(value="-")
            ttk.Label(card, textvariable=var, font=("Segoe UI", 14, "bold")).pack(anchor="center")
            self.card_vars[titulo] = var

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=12, pady=12)
        self.aba_ocup = ttk.Frame(self.notebook)
        self.aba_renda = ttk.Frame(self.notebook)
        self.aba_rank = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_ocup, text="Ocupações")
        self.notebook.add(self.aba_renda, text="Renda")
        self.notebook.add(self.aba_rank, text="Ranking")

        self.fig_ocup = Figure(figsize=(9, 4.8), dpi=100)
        self.ax_ocup = self.fig_ocup.add_subplot(111)
        self.canvas_ocup = FigureCanvasTkAgg(self.fig_ocup, master=self.aba_ocup)
        self.canvas_ocup.get_tk_widget().pack(fill="both", expand=True)

        self.fig_renda = Figure(figsize=(9, 4.8), dpi=100)
        self.ax_renda = self.fig_renda.add_subplot(111)
        self.canvas_renda = FigureCanvasTkAgg(self.fig_renda, master=self.aba_renda)
        self.canvas_renda.get_tk_widget().pack(fill="both", expand=True)

        self.tree = ttk.Treeview(self.aba_rank, columns=("ra", "media", "pessoas", "pct"), show="headings")
        for col, nome in [("ra", "RA/localidade"), ("media", "Renda média"), ("pessoas", "Pessoas"), ("pct", "% do filtro")]:
            self.tree.heading(col, text=nome)
            self.tree.column(col, width=260 if col == "ra" else 110, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=8, pady=8)

    def combo(self, pai, texto, var, valores, coluna):
        """Cria um rótulo e uma caixa de seleção."""
        ttk.Label(pai, text=texto).grid(row=0, column=coluna * 2, sticky="w", padx=(0, 4))
        cb = ttk.Combobox(pai, textvariable=var, values=valores, state="readonly", width=24)
        cb.grid(row=0, column=coluna * 2 + 1, padx=(0, 8))
        cb.bind("<<ComboboxSelected>>", lambda e: self.atualizar())
        return cb

    def opcoes(self, coluna, primeira):
        """Monta lista de opções dos filtros."""
        vals = self.ocupados[coluna].dropna().unique().tolist()
        vals = sorted(vals, key=lambda x: int(str(x).split(" - ")[0]) if str(x).split(" - ")[0].isdigit() else 99999)
        return [primeira] + vals

    def aplicar_filtros(self):
        """Filtra a base conforme os widgets selecionados."""
        df = self.ocupados.copy()
        if self.var_ra.get() != "Todas":
            df = df[df["ra"] == self.var_ra.get()]
        if self.var_genero.get() != "Todos":
            df = df[df["genero"] == self.var_genero.get()]
        if self.var_ensino.get() != "Todos":
            df = df[df["ensino"] == self.var_ensino.get()]
        renda_min = float(self.var_renda.get())
        df = df[df["renda"].fillna(0) >= renda_min]
        self.filtrado = df

    def atualizar_cards(self):
        """Atualiza os cards antigos de estatísticas."""
        st = estatisticas(self.filtrado)
        total = st["ocupados"]
        com_renda = st["com_renda"]
        pct = (com_renda / total * 100) if total else 0
        self.card_vars["Ocupados filtrados"].set(f"{total:,}".replace(",", "."))
        self.card_vars["Com renda válida"].set(f"{com_renda:,} ({pct:.1f}%)".replace(",", "."))
        self.card_vars["Média da renda"].set(moeda(st["media"]))
        self.card_vars["Mediana da renda"].set(moeda(st["mediana"]))

    def atualizar(self):
        """Atualiza estatísticas, gráficos e ranking."""
        self.aplicar_filtros()
        self.atualizar_cards()
        self.grafico_ocupacoes()
        self.grafico_renda()
        self.atualizar_ranking()

    def grafico_ocupacoes(self):
        """Desenha gráfico de barras com quantidade e porcentagem por ocupação."""
        self.ax_ocup.clear()
        cont = self.filtrado["ocupacao"].value_counts().head(10)
        total = cont.sum()
        if cont.empty:
            self.ax_ocup.text(0.5, 0.5, "Sem dados", ha="center", va="center")
        else:
            cont = cont.iloc[::-1]
            barras = self.ax_ocup.barh(cont.index, cont.values)
            for barra, valor in zip(barras, cont.values):
                pct = valor / total * 100 if total else 0
                self.ax_ocup.text(barra.get_width(), barra.get_y() + barra.get_height() / 2, f" {valor} ({pct:.1f}%)", va="center")
            self.ax_ocup.set_title("Distribuição das ocupações")
            self.ax_ocup.set_xlabel("Quantidade de pessoas")
        self.fig_ocup.tight_layout()
        self.canvas_ocup.draw()

    def grafico_renda(self):
        """Desenha histograma da renda do trabalho principal."""
        self.ax_renda.clear()
        renda = self.filtrado["renda"].dropna()
        if renda.empty:
            self.ax_renda.text(0.5, 0.5, "Sem renda válida", ha="center", va="center")
        else:
            limite = renda.quantile(0.95)
            self.ax_renda.hist(renda[renda <= limite], bins=25)
            self.ax_renda.set_title("Distribuição da renda do trabalho principal")
            self.ax_renda.set_xlabel("Renda em R$")
            self.ax_renda.set_ylabel("Pessoas")
        self.fig_renda.tight_layout()
        self.canvas_renda.draw()

    def atualizar_ranking(self):
        """Atualiza ranking das localidades por renda média."""
        for item in self.tree.get_children():
            self.tree.delete(item)
        base = self.filtrado.dropna(subset=["renda"])
        total = len(base)
        ranking = []
        for ra, grupo in base.groupby("ra"):
            if len(grupo) >= 5:
                ranking.append((ra, grupo["renda"].mean(), len(grupo), len(grupo) / total * 100 if total else 0))
        for ra, media, pessoas, pct in ordenar_manual(ranking, indice=1, decrescente=True)[:15]:
            self.tree.insert("", "end", values=(ra, moeda(media), pessoas, f"{pct:.1f}%"))

    def gerar_relatorio(self):
        """Gera um relatório simples em TXT com os dados do filtro atual."""
        caminho = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Arquivo de texto", "*.txt")]
        )

        if not caminho:
            return

        st = estatisticas(self.filtrado)
        total = st["ocupados"]
        com_renda = st["com_renda"]
        pct_renda = (com_renda / total * 100) if total else 0

        filtros_usados = [
            f"Localidade: {self.var_ra.get()}",
            f"Gênero: {self.var_genero.get()}",
            f"Escolaridade: {self.var_ensino.get()}",
            f"Renda mínima: {moeda(float(self.var_renda.get()))}",
        ]

        ocupacoes = self.filtrado["ocupacao"].value_counts().head(10)

        base_ranking = self.filtrado.dropna(subset=["renda"])
        ranking = []
        total_ranking = len(base_ranking)
        for ra, grupo in base_ranking.groupby("ra"):
            if len(grupo) >= 5:
                ranking.append((
                    ra,
                    grupo["renda"].mean(),
                    len(grupo),
                    len(grupo) / total_ranking * 100 if total_ranking else 0
                ))
        ranking = ordenar_manual(ranking, indice=1, decrescente=True)[:10]

        with open(caminho, "w", encoding="utf-8") as arquivo:
            arquivo.write("RELATÓRIO SIMPLES — PDAD 2024\n")
            arquivo.write("Recorte E: Trabalho e Ocupação\n")
            arquivo.write("=" * 45 + "\n\n")

            arquivo.write("FILTROS APLICADOS\n")
            arquivo.write("-" * 45 + "\n")
            for item in filtros_usados:
                arquivo.write(item + "\n")

            arquivo.write("\nRESUMO DO FILTRO\n")
            arquivo.write("-" * 45 + "\n")
            arquivo.write(f"Ocupados filtrados: {total}\n")
            arquivo.write(f"Pessoas com renda válida: {com_renda} ({pct_renda:.1f}%)\n")
            arquivo.write(f"Média da renda: {moeda(st['media'])}\n")
            arquivo.write(f"Mediana da renda: {moeda(st['mediana'])}\n")

            arquivo.write("\nPRINCIPAIS OCUPAÇÕES\n")
            arquivo.write("-" * 45 + "\n")
            if ocupacoes.empty:
                arquivo.write("Sem dados para ocupações.\n")
            else:
                soma_ocupacoes = ocupacoes.sum()
                for ocupacao, quantidade in ocupacoes.items():
                    pct = quantidade / soma_ocupacoes * 100 if soma_ocupacoes else 0
                    arquivo.write(f"{ocupacao}: {quantidade} pessoas ({pct:.1f}%)\n")

            arquivo.write("\nRANKING DE LOCALIDADES POR RENDA MÉDIA\n")
            arquivo.write("-" * 45 + "\n")
            if not ranking:
                arquivo.write("Sem dados suficientes para montar o ranking.\n")
            else:
                for posicao, (ra, media, pessoas, pct) in enumerate(ranking, start=1):
                    arquivo.write(
                        f"{posicao}. {ra} | média: {moeda(media)} | "
                        f"pessoas: {pessoas} | {pct:.1f}% do filtro\n"
                    )

        messagebox.showinfo("Relatório gerado", "Relatório TXT salvo com sucesso!")




def centralizar(janela, largura=430, altura=170):
    """Centraliza uma janela na tela."""
    janela.update_idletasks()
    x = (janela.winfo_screenwidth() // 2) - (largura // 2)
    y = (janela.winfo_screenheight() // 2) - (altura // 2)
    janela.geometry(f"{largura}x{altura}+{x}+{y}")


def carregar_com_tela(root):
    """Mostra uma tela de carregamento enquanto os arquivos são lidos."""
    tela = tk.Toplevel(root)
    tela.title("Carregando dados")
    tela.resizable(False, False)
    tela.protocol("WM_DELETE_WINDOW", lambda: None)
    centralizar(tela)

    caixa = ttk.Frame(tela, padding=18)
    caixa.pack(fill="both", expand=True)
    ttk.Label(caixa, text="Carregando PDAD 2024", font=("Segoe UI", 14, "bold")).pack(anchor="w")
    status = tk.StringVar(value="Preparando leitura dos arquivos...")
    ttk.Label(caixa, textvariable=status).pack(anchor="w", pady=(8, 6))
    barra = ttk.Progressbar(caixa, mode="determinate", maximum=100)
    barra.pack(fill="x", pady=(2, 8))
    porcentagem = tk.StringVar(value="0%")
    ttk.Label(caixa, textvariable=porcentagem, font=("Segoe UI", 9)).pack(anchor="e")

    def progresso(texto, valor):
        status.set(texto)
        barra["value"] = valor
        porcentagem.set(f"{int(valor)}%")
        tela.update_idletasks()
        tela.update()

    try:
        dados = carregar_dados(progresso)
    finally:
        tela.destroy()
    return dados


def main():
    """Inicia o sistema com tela de carregamento."""
    root = tk.Tk()
    root.withdraw()
    try:
        moradores, ocupados, total_dom = carregar_com_tela(root)
        root.deiconify()
        SistemaPDAD(root, moradores, ocupados, total_dom)
        root.mainloop()
    except Exception as erro:
        root.deiconify()
        messagebox.showerror("Erro", str(erro))
        root.destroy()


if __name__ == "__main__":
    main()
